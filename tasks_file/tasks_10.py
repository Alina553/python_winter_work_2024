"""Задача 10-1
Есть текстовый файл 'test1.txt'. Надо прочитать содержимое этого файла и создать файл 'test2.txt'. В нем должны быть строчки из первого файла, но в обратном порядке. В каждой строке должен быть поменян порядок слов на противоположный.
Например, исходный файл:
Мой дядя самых честных правил Когда не в шутку занемог
Результат:
занемог шутку в не Когда
правил честных самых дядя Мой"""

with open('../test.txt', 'r') as file1:
    lines = file1.readlines()
    print(lines)

with open('../test1.txt', 'w') as file2:
    for line in lines[::-1]:
        word = line.split()
        reversed_line = ' '.join(word[::-1])
        file2.write(reversed_line + '\n')

with open('../test1.txt', 'r') as file:
    lines = file.readlines()
    print(lines)


"""Задача 10-2
Дан эксельный файл. На странице находится список сотрудников и их ЗП.
Надо создать еще одну страницу в этом файле и поместить туда отсортированный список фамилий и их ЗП,
а в последней строчке поместить слово ИТОГО: и сумму всех ЗП. Сортировка по убыванию ЗП сотрудников.
Например, исходная страница файла:
Сидоров 100 Петров 200 Иванов 300
Результат:
Иванов 300 Петров 200 Сидоров 100 ИТОГО: 600"""

import openpyxl

wb = openpyxl.load_workbook('test.xlsx') # загружаем файл Excel
sheet = wb['Sheet'] # выбираем страницу, на которой находится список сотрудников и их ЗП

data = [] #создаем пустой сисок для хранения данных

for row in sheet.iter_rows(values_only=True):
    data.append(row)

print(data)
sorted_data = sorted(data, key= lambda x: x[1], reverse=True) #сортируем список по убыванию зп сотрудников x[1] значит выбираем 2 элемент
print(sorted_data)

new_sheet = wb.create_sheet('Sorted Employees')

row_num = 1
for item in sorted_data:
    a = new_sheet.cell(row = row_num, column =1, value = item[0]) #записываем фамилию
    b = new_sheet.cell(row = row_num, column =2, value = item[1]) # записываем зп
    row_num += 1
    print(a.value,b.value)

# находим сумму зп
sum = 0
for i in sorted_data:
    sum += i[1]
    print(sum)

new_sheet.cell(row = row_num, column=1, value='Итого:')
new_sheet.cell(row = row_num, column=2, value=sum)

wb.save('test.xlsx')

#Проверяем
for row in new_sheet.iter_rows(values_only=True):
    print(row)

"""Задача 10-3
Дан исходный эксельный файл со списком людей и их ЗП.
Следует создать еще одну страницу со статистическими данными исходного списка. Например, исходная страница:
Сидоров 100
Петров 200
Иванов 300
Федоров 1000
Результат:
Максимальное значение 1000 Минимальное значение 100 Сумма 1600 Среднее арифметическое 800 Медиана 250"""

import openpyxl
import statistics

wb = openpyxl.load_workbook('../tasks_file/test.xlsx') # загружаем файл Excel

print(wb.sheetnames)
sheet = wb['Sheet']

data = []

for row in sheet.iter_rows(values_only=True):
    data.append(row[1])
    print(data)

a = max(data)
b = min(data)
s = sum(data)
m = statistics.median(data)

print(a,b,s,m)

data_new = [('Максимальное значение',a),('Минимальное значение',b),('Сумма',s),('Медиана',m)]
new_sheet = wb.create_sheet('Static data')

#добавлям новые строки в новую вкладку
for i, j in enumerate(data_new, start=1):
    new_sheet.cell(row = i, column = 1, value = j[0])
    new_sheet.cell(row = i, column = 2, value = j[1])

#проверяем
for item in new_sheet.iter_rows(values_only=True):
    print(item)

wb.save('../tasks_file/test.xlsx')


