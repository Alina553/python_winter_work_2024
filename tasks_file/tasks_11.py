"""Задача 11-1
Каждый третий четверг каждого месяца билеты в Эрмитаж бесплатны.
Напечатайте список дат в 2024 году, когда вход бесплатен."""

import datetime

start_date = datetime.date(2024,1,1)
end_date= datetime.date(2024,12,31)
print(start_date.day)


free_dates = []

current_date = start_date
while current_date <= end_date:
    if current_date.weekday() == 3 and (current_date.day // 7) % 3 == 0:
        free_dates.append(current_date)
    current_date += datetime.timedelta(days=1)

for date in free_dates:
    print(date)


"""Задача 11-2
Дан файл с расширением .txt, содержащий в каждой строчке следующую информацию: номер, фамилия, имя, компания, зарплата, разделенные запятыми.
Создайте Эксельный файл, в который перенесите эту информацию, предварительно отсортировав этот список по компании, по фамилии и по имени.
В конце списка добавьте строку: ИТОГО и суммарное значение всех зарплат."""

import openpyxl

# формируем список из подсписков (стрипом отбрасываем пробелы слева/справа, сплитом разделяем строку запятой и образуем список
with open('test_11.txt', 'r') as f:
    lst = []
    for i in f:
        lst.append(i.strip().split(
            ','))  # [['1', ' Иванов', ' Иван', ' Ромашка', ' 12000'], ['2', ' Петров', ' Кирилл', ' Ирис', ' 13400'], ['3', ' Казаков', ' Олег', ' Бутон', ' 77600']]

print(lst)

# создаем в эксель страницу sheet есл и ее небыло
wb = openpyxl.load_workbook('test_11.xlsx')
if 'Sheet' not in wb.sheetnames:
    wb.create_sheet('Sheet')
ws = wb['Sheet']
for i in range(1, ws.max_row + 1):
    for j in range(1, ws.max_column + 1):
        ws.cell(i, j).value = None  # вдруг есть информация поэтому затерли все данные Noneом

# сортируем список lst с помощью лямбды по компании6 фамилии, имени
lst_sorted = sorted(lst, key=lambda x: (x[3], x[1], x[2]))
print(lst_sorted)

# перебирем lst_sorted начиная с 1
total = 0
for i, j in enumerate(lst_sorted, 1):
    for k in range(5):  #
        if j[k].isdigit():  # если значение ячейки цифра, то пусть тип будет интенжер
            ws.cell(i, k + 1).value = int(j[k])
        else:
            ws.cell(i, k + 1).value = j[k]

    total += int(j[4])  # сумма зарплат int тк в списке формат зп был строка
print(total)

ws.cell(i + 1, 2).value = 'Итого'
ws.cell(i + 1, 5).value = total

wb.save('test_11.xlsx')

"""Задача 11-3
Напишите функцию, которая переводит арабские числа в римские. Например: 2023 ->MMXXIII"""

d = {1000:'M',900:'CM',500:'D',400:'CD',100:'C',90:'CX',50:'L',40:'XL',10:'X',9:'IX',5:'V',4:'IV',1:'I'}

def arab_to_rim(x):
    result = ''
    while x > 0:
        for i in d:
            if x >= i:
                result += d[i]
                x -= i
                break
    return result

print(arab_to_rim(int(input())))

def arab_to_rim2(x):
    result = ''
    for i in d:
        while x > 0 and x >=i:
            result += d[i]
            x -= i
    return result

print(arab_to_rim2(int(input())))
